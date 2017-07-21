from PyQt5 import QtGui, QtWidgets, QtCore 
from PyQt5.QtWidgets import QWidget, QApplication, QPushButton, QHBoxLayout, QRadioButton, QButtonGroup, QLabel, QLineEdit, QFormLayout, QWidget, QMessageBox
from PyQt5.QtGui import QIcon, QPainter, QColor, QPen, QImage, QPalette, QBrush, QPixmap
from PyQt5.QtCore import QSize, Qt, QThread, pyqtSignal, pyqtSlot
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import sys
import time
import os

def trap_exc_during_debug(*args):
    print(args)

sys.excepthook = trap_exc_during_debug

class workerThread(QThread):

    signal = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.abort = False
        
    @pyqtSlot()
    def run(self):
        #print("yeah yeah ")
        time.sleep(0.1)
        app.processEvents()
        self.signal.emit('Done')
        
    def __del__(self):
        #print("okay okay")
        self.abort = True
        self.wait()
        
class General(QWidget):
    
    def __init__(self):
        
        super().__init__()
        self.initUI()
        
    def initUI(self):
        
        saveButton = QPushButton("SAVE",self)
        clearButton = QPushButton("CLEAR",self)
        saveButton.setFont(QtGui.QFont("Calibri", 13))
        clearButton.setFont(QtGui.QFont("Calibri", 13))

        saveButton.move(100,680)
        clearButton.move(260,680)
        
        comboBoxyopass = QtWidgets.QComboBox(self)
        comboBoxyopass.addItem("SELECT")
        i=2050
        while i>=2000:
            comboBoxyopass.addItem(str(i))
            i-=1
        comboBoxyopass.setMinimumHeight(35)
        comboBoxyopass.setFixedWidth(150)
        comboBoxyopass.setFont(QtGui.QFont("Calibri", 14))

        comboBoxtrcentr = QtWidgets.QComboBox(self)
        comboBoxtrcentr.addItem("SELECT CENTER")
        comboBoxtrcentr.addItem('Jaipur   ')
        comboBoxtrcentr.addItem('Hyderabad')
        comboBoxtrcentr.addItem('Raipur   ')
        comboBoxtrcentr.addItem('Lucknow  ')
        comboBoxtrcentr.addItem('Pune     ')
        comboBoxtrcentr.addItem('Vizag    ')
        comboBoxtrcentr.addItem('Bhopal   ')
        comboBoxtrcentr.addItem('Delhi    ')
        comboBoxtrcentr.setMinimumHeight(35)
        comboBoxtrcentr.setFixedWidth(180)
        comboBoxtrcentr.setFont(QtGui.QFont("Calibri", 14))

        comboBoxcourse = QtWidgets.QComboBox(self)
        comboBoxcourse.addItem("SELECT COURSE")
        comboBoxcourse.addItem('ESR    (30 Days)     ')
        comboBoxcourse.addItem('ESR    (45 Days)     ')
        comboBoxcourse.addItem('Matlab   (30 Days)   ')
        comboBoxcourse.addItem('IOT   (15 Days)      ')
        comboBoxcourse.addItem('IOT   (30 Days)      ')
        comboBoxcourse.addItem('JAVA   (30 Days)     ')
        comboBoxcourse.addItem('Python   (30 Days)   ')
        comboBoxcourse.addItem('PLC-SCADA   (30 Days)')
        comboBoxcourse.addItem('C/C++   (45 Days)    ')
        comboBoxcourse.addItem('Android   (15 Days)  ')
        comboBoxcourse.addItem('Android   (30 Days)  ')
        comboBoxcourse.setMinimumHeight(35)
        comboBoxcourse.setFixedWidth(200)
        comboBoxcourse.setFont(QtGui.QFont("Calibri", 14))

        comboBoxsem = QtWidgets.QComboBox(self)
        comboBoxsem.addItem("SELECT")
        i=1
        while i<=8:
            comboBoxsem.addItem(str(i))
            i+=1
        comboBoxsem.addItem("Passed Out")
        comboBoxsem.setMinimumHeight(35)
        comboBoxsem.setFixedWidth(100)
        comboBoxsem.setFont(QtGui.QFont("Calibri", 14))

        comboBoxstate = QtWidgets.QComboBox(self)
        comboBoxstate.addItem("SELECT")
        comboBoxstate.addItem('Andhra Pradesh')
        comboBoxstate.addItem('Arunachal Pradesh')
        comboBoxstate.addItem('Assam')
        comboBoxstate.addItem('Bihar')
        comboBoxstate.addItem('Goa')
        comboBoxstate.addItem('Gujarat')
        comboBoxstate.addItem('Haryana')
        comboBoxstate.addItem('Himachal Pradesh')
        comboBoxstate.addItem('Jammu & Kashmir')
        comboBoxstate.addItem('Karnataka')
        comboBoxstate.addItem('Kerala')
        comboBoxstate.addItem('Madhya Pradesh')
        comboBoxstate.addItem('Maharashtra')
        comboBoxstate.addItem('Manipur')
        comboBoxstate.addItem('Meghalaya')
        comboBoxstate.addItem('Mizoram')
        comboBoxstate.addItem('Nagaland')
        comboBoxstate.addItem('Orissa')
        comboBoxstate.addItem('Punjab')
        comboBoxstate.addItem('Rajasthan')
        comboBoxstate.addItem('Sikkim')
        comboBoxstate.addItem('Tamil Nadu')
        comboBoxstate.addItem('Tripura')
        comboBoxstate.addItem('Uttar Pradesh')
        comboBoxstate.addItem('West Bengal')
        comboBoxstate.addItem('Chhattisgarh')
        comboBoxstate.addItem('Uttarakhand')
        comboBoxstate.addItem('Jharkhand')
        comboBoxstate.addItem('Telangana')
        comboBoxstate.setMinimumHeight(35)
        comboBoxstate.setFixedWidth(250)
        comboBoxstate.setFont(QtGui.QFont("Calibri", 14))

        hboxsex = QHBoxLayout()
        hboxsex.setSpacing(60)
        r1 = QRadioButton("Male")
        r1.setFont(QtGui.QFont("Calibri", 10.5, QtGui.QFont.Bold))
        r1.setMinimumHeight(30)
        r2 = QRadioButton("Female")
        r2.setFont(QtGui.QFont("Calibri", 10.5, QtGui.QFont.Bold))
        r2.setMinimumHeight(30)
        widgetsex=QWidget(self)
        groupsex=QButtonGroup(widgetsex)
        groupsex.addButton(r1)
        groupsex.addButton(r2)
        hboxsex.addWidget(r1)
        hboxsex.addWidget(r2)
        hboxsex.addStretch()
        
        headerfont = QtGui.QFont("Cambria", 13, QtGui.QFont.Bold)
        saveloc=str("Student_List.xlsx")

        l1 = QLabel("Name: ")
        l1.setFont(headerfont)
        l1.setMinimumHeight(30)
        l1.setFixedWidth(180)
        text1 = QLineEdit()
        text1.setFixedWidth(600)
        text1.setMinimumHeight(30)
        text1.setFont(QtGui.QFont("Times", 11))
        

        l2 = QLabel("Email Id: ")
        l2.setFont(headerfont)
        l2.setMinimumHeight(30)
        l2.setFixedWidth(180)
        text2 = QLineEdit()
        text2.setFixedWidth(600)
        text2.setMinimumHeight(30)
        text2.setFont(QtGui.QFont("Times", 11))

        l3 = QLabel("Contact No.: ")
        l3.setFont(headerfont)
        l3.setMinimumHeight(30)
        l3.setFixedWidth(180)
        text3 = QLineEdit()
        text3.setFixedWidth(600)
        text3.setMinimumHeight(30)
        text3.setFont(QtGui.QFont("Times", 11))
        
        l4 = QLabel("City: ")
        l4.setFont(headerfont)
        l4.setMinimumHeight(30)
        l4.setFixedWidth(180)
        text4 = QLineEdit()
        text4.setFixedWidth(600)
        text4.setMinimumHeight(30)
        text4.setFont(QtGui.QFont("Times", 11))

        l5 = QLabel("State: ")
        l5.setFont(headerfont)
        l5.setMinimumHeight(30)
        l5.setFixedWidth(180)

        l6 = QLabel("College: ")
        l6.setFont(headerfont)
        l6.setMinimumHeight(30)
        l6.setFixedWidth(180)
        text6 = QLineEdit()
        text6.setFixedWidth(600)
        text6.setMinimumHeight(30)
        text6.setFont(QtGui.QFont("Times", 11))
        
        l7 = QLabel("Branch: ")
        l7.setFont(headerfont)
        l7.setMinimumHeight(30)
        l7.setFixedWidth(180)
        text7 = QLineEdit()
        text7.setFixedWidth(600)
        text7.setMinimumHeight(30)
        text7.setFont(QtGui.QFont("Times", 11))

        l8 = QLabel("Semester: ")
        l8.setFont(headerfont)
        l8.setMinimumHeight(30)
        l8.setFixedWidth(180)

        l9 = QLabel("Year Of Passing: ")
        l9.setFont(headerfont)
        l9.setFixedWidth(180)

        l10 = QLabel("Course: ")
        l10.setFont(headerfont)
        l10.setMinimumHeight(30)
        l10.setFixedWidth(180)

        l11 = QLabel("Batch: ")
        l11.setFont(headerfont)
        l11.setMinimumHeight(30)
        l11.setFixedWidth(180)
        text11 = QLineEdit()
        text11.setFixedWidth(600)
        text11.setMinimumHeight(30)
        text11.setFont(QtGui.QFont("Times", 11))

        l12 = QLabel("Training Center: ")
        l12.setFont(headerfont)
        l12.setMinimumHeight(30)
        l12.setFixedWidth(180)

        l13 = QLabel("SEX: ")
        l13.setFont(headerfont)
        l13.setFixedWidth(180)

        l14 = QLabel("Save File As: ")
        l14.setFont(headerfont)
        l14.setMinimumHeight(30)
        l14.setFixedWidth(180)
        text14 = QLineEdit()
        text14.setFixedWidth(600)
        text14.setMinimumHeight(30)
        text14.setFont(QtGui.QFont("Times", 11,QtGui.QFont.Bold))
        text14.setText(saveloc)

        l15 = QLabel("Query/Regarding What: ")
        l15.setFont(QtGui.QFont("Cambria", 12, QtGui.QFont.Bold))
        l15.setMinimumHeight(30)
        l15.setFixedWidth(200)
        text15 = QLineEdit()
        text15.setFixedWidth(600)
        text15.setMinimumHeight(30)
        text15.setFont(QtGui.QFont("Times", 11))

        hboxcourse = QHBoxLayout()
        hboxcourse.setSpacing(25)
        l16 = QLabel("Others: ")
        l16.setFont(headerfont)
        l16.setMinimumHeight(30)
        l16.setFixedWidth(100)
        text16 = QLineEdit()
        text16.setFixedWidth(250)
        text16.setMinimumHeight(30)
        text16.setFont(QtGui.QFont("Times", 11))
        hboxcourse.addWidget(comboBoxcourse)
        hboxcourse.addWidget(l16)
        hboxcourse.addWidget(text16)
        hboxcourse.addStretch()

        hboxstate = QHBoxLayout()
        hboxstate.setSpacing(25)
        l17 = QLabel("Others: ")
        l17.setFont(headerfont)
        l17.setMinimumHeight(30)
        l17.setFixedWidth(70)
        text17 = QLineEdit()
        text17.setFixedWidth(230)
        text17.setMinimumHeight(30)
        text17.setFont(QtGui.QFont("Times", 11))
        hboxstate.addWidget(comboBoxstate)
        hboxstate.addWidget(l17)
        hboxstate.addWidget(text17)
        hboxstate.addStretch()
        
        fbox = QFormLayout()
        fbox.setVerticalSpacing(10)
        
        fbox.addRow(l1,text1)
        fbox.addRow(l2,text2)
        fbox.addRow(l3,text3)
        fbox.addRow(l4,text4)
        fbox.addRow(l5,hboxstate)
        fbox.addRow(l6,text6)
        fbox.addRow(l7,text7)
        fbox.addRow(l8,comboBoxsem)
        fbox.addRow(l9,comboBoxyopass)
        fbox.addRow(l10,hboxcourse)
        fbox.addRow(l11,text11)
        fbox.addRow(l12,comboBoxtrcentr)

        l18 = QLabel("Training Session: ")
        l18.setFont(headerfont)
        l18.setMinimumHeight(30)
        l18.setFixedWidth(200)

        hboxperiod = QHBoxLayout()
        hboxperiod.setSpacing(70)
        r3 = QRadioButton("Summer Training")
        r3.setFont(QtGui.QFont("Calibri", 10, QtGui.QFont.Bold))
        r3.setMinimumHeight(30)
        r4 = QRadioButton("Winter Training")
        r4.setFont(QtGui.QFont("Calibri", 10, QtGui.QFont.Bold))
        r4.setMinimumHeight(30)
        r5 = QRadioButton("Project Based")
        r5.setFont(QtGui.QFont("Calibri", 10, QtGui.QFont.Bold))
        r5.setMinimumHeight(30)
        r6 = QRadioButton("Other")
        r6.setFont(QtGui.QFont("Calibri", 10, QtGui.QFont.Bold))
        r6.setMinimumHeight(30)
        widgetperiod=QWidget(self)
        groupperiod=QButtonGroup(widgetperiod)
        groupperiod.addButton(r3)
        groupperiod.addButton(r4)
        groupperiod.addButton(r5)
        groupperiod.addButton(r6)
        hboxperiod.addWidget(r3)
        hboxperiod.addWidget(r4)
        hboxperiod.addWidget(r5)
        hboxperiod.addWidget(r6)
        hboxperiod.addStretch()
        fbox.addRow(l18,hboxperiod)
        fbox.addRow(l13,hboxsex)
        fbox.addRow(l15,text15)
        fbox.addRow(l14,text14)        
        
        self.lineedits = [text1,text2,text3,text4,text6,text7,text11,text14,text15,text16,text17]
        self.saveedit=[text14]
        self.comboBox = [comboBoxstate,comboBoxsem,comboBoxyopass,comboBoxtrcentr,comboBoxcourse]
        self.radiobutton=[r1,r2,r3,r4,r5,r6]
        saveButton.clicked.connect(self.saveClicked)
        clearButton.clicked.connect(self.clearClicked)
        
        self.setLayout(fbox)
        
        try:
            self.setWindowState(QtCore.Qt.WindowMaximized)
        except:
            self.setGeometry(10, 30, 1350, 750)
            
        self.setWindowTitle('Managemet System Software ')
        self.setWindowIcon(QIcon('logso.png'))                          # Enter your Icon Image url here 
        oImage = QImage("image2.jpg")                                   # Enter your Background Image url here
        sImage = oImage.scaled(QSize(1350,750))                   
        palette = QPalette()
        palette.setBrush(10, QBrush(sImage))
        self.setPalette(palette)
        self.show()
        
    def validContact(self,phone_number):
        if len(phone_number)!=10:
            return False
        else:
            for i in range(10):
                if (ord(phone_number[i])-48) not in range(10):
                    return False
        return True
    
    def savesuccess(self):
        savemsg = QMessageBox()
        savemsg.setIcon(QMessageBox.Information)
        savemsg.setText("Your Enteries Have Been Saved successfully !")
        savemsg.setWindowTitle("SAVED")
        savemsg.setWindowIcon(QIcon('logso.png'))
        savemsg.exec()
        
    def errorcontactmsg(self):
        erroremsg = QMessageBox()
        erroremsg.setIcon(QMessageBox.Warning)
        erroremsg.setText("Please Enter A Valid Email ID!!")
        erroremsg.setWindowTitle("Invalid Email ID")
        erroremsg.setWindowIcon(QIcon('logso.png'))
        erroremsg.exec()
        
    def erroremailmsg(self):
        erroremsg = QMessageBox()
        erroremsg.setIcon(QMessageBox.Warning)
        erroremsg.setText("Please Enter A Valid Email ID!!")
        erroremsg.setWindowTitle("Invalid Email ID")
        erroremsg.setWindowIcon(QIcon('logso.png'))
        erroremsg.exec()
                    
    def errornormmsg(self):
        errormsg = QMessageBox()
        errormsg.setIcon(QMessageBox.Warning)
        errormsg.setText("Please Enter The necessary Fields: Name, Email Id, Conact No. !!")
        errormsg.setWindowTitle("Error Report")
        errormsg.setWindowIcon(QIcon('logso.png'))
        errormsg.exec()
                
    def crashingmsg(self):
        crashmsg = QMessageBox()
        crashmsg.setIcon(QMessageBox.Critical)
        crashmsg.setText("The XML File is Already Open")
        crashmsg.setDetailedText("Please Close the XML file and                   Try again")
        crashmsg.setWindowTitle("Programe Crashing")
        crashmsg.setWindowIcon(QIcon('logso.png'))
        crashmsg.exec()
        
    def saveClicked(self):
        try:
            tex=[]
            for edit in self.lineedits:
                tex.append(str(edit.text()))
            if tex[0] and tex[1] and tex[2]:
                if ('@gmail.com' in str(tex[1])) or ('@yahoo.com' in str(tex[1])) or ('@' in str(tex[1])):
                    contact=str(tex[2])
                    if self.validContact(contact):
                        try :
                            wb = load_workbook(os.path.join("C:\Management_sys_excels",str(tex[7])))
                            ws = wb.active
                            row=ws._current_row
                            sno=int(ws['A'+str(row)].internal_value)
                            tkno=int(ws['B'+str(row)].internal_value)
                            row+=1
                        except:
                            wb = Workbook()
                            ws = wb.active
                            ws.title="Student_List"
                            ft = Font(bold=True)
                            ws['A1']="S.NO"
                            ws['B1']="TOKEN NO."
                            ws['C1']="NAME"
                            ws['D1']="EMAIL ID"
                            ws['E1']="CONTACT NO."
                            ws['F1']="CITY"
                            ws['G1']="STATE"
                            ws['H1']="COLLEGE"
                            ws['I1']="BRANCH"
                            ws['J1']="SEMESTER"
                            ws['K1']="YEAR OF PASSING"
                            ws['L1']="COURSE"
                            ws['M1']="DATE OF REGISTERATION"
                            ws['N1']="BATCH"
                            ws['O1']="TRAINING CENTER"
                            ws['P1']="TRAINING SESSION"
                            ws['Q1']="SEX"
                            ws['R1']="QUERY/REGARDING WHAT"
                            ws['S1']="TOTAL FEE"
                            ws['T1']="SUBMITTED FEE"
                            ws['U1']="REMAINING FEE"
                            ws['V1']="DISCOUNT"
                            ws['W1']="REMARKS"
                            ws['X1']="ADDITIONAL REMARKS"
                            ws['Y1']="SYSTEM DATE AND TIME"
                            j=65
                            while j<=88:
                                ws[chr(j)+'1'].font=ft
                                j+=1
                            sno=0
                            tkno=0
                            row=2
                        ws['A'+str(row)]=int(sno+1)
                        ws['B'+str(row)]=int(tkno+1)

                        j=0
                        while j<4:
                            ws[chr(67+j)+str(row)]=str(tex[j])
                            j+=1
                            
                        stateopt=str(self.comboBox[0].currentText())
                        if stateopt != "SELECT":
                            ws['G'+str(row)]=stateopt
                        else:
                            if tex[10]:
                                ws['G'+str(row)]=str(tex[10])
                                
                        ws['H'+str(row)]=str(tex[4])
                        ws['I'+str(row)]=str(tex[5])
                        ws['J'+str(row)]=str(self.comboBox[1].currentText())
                        ws['K'+str(row)]=str(self.comboBox[2].currentText())

                        courseopt=str(self.comboBox[4].currentText())
                        if courseopt != "SELECT COURSE":
                            ws['L'+str(row)]=courseopt
                        else:
                            if tex[9]:
                                ws['L'+str(row)]=str(tex[9])
                                
                        ws['N'+str(row)]=str(tex[6])
                        ws['O'+str(row)]=str(self.comboBox[3].currentText())
                        
                        if self.radiobutton[2].isChecked():
                            ws['P'+str(row)]=str(self.radiobutton[2].text())
                        else:
                            if self.radiobutton[3].isChecked():
                                ws['P'+str(row)]=str(self.radiobutton[3].text())
                            else:
                                if self.radiobutton[4].isChecked():
                                    ws['P'+str(row)]=str(self.radiobutton[4].text())
                                else:
                                    if self.radiobutton[5].isChecked():
                                        ws['P'+str(row)]=str(self.radiobutton[5].text())
                            
                        if self.radiobutton[0].isChecked():
                            ws['Q'+str(row)]=str(self.radiobutton[0].text())
                        else:
                            if self.radiobutton[1].isChecked():
                                ws['Q'+str(row)]=str(self.radiobutton[1].text())
                                
                        ws['R'+str(row)]=str(tex[8])
                        ws['Y'+str(row)]=str(QtCore.QDateTime.currentDateTime().toString())
                        if not os.path.exists("C:\Management_sys_excels"):
                            os.makedirs("C:\Management_sys_excels")
                            
                        wb.save(os.path.join("C:\Management_sys_excels",str(tex[7])))

                        self.workerthread = workerThread()
                        self.workerthread.signal.connect(self.savesuccess)
                        self.workerthread.start()
                    else:
                        self.workerthread = workerThread()
                        self.workerthread.signal.connect(self.errorcontactmsg)
                        self.workerthread.start()
                else:
                    self.workerthread = workerThread()
                    self.workerthread.signal.connect(self.erroremailmsg)
                    self.workerthread.start()
            else:
                self.workerthread = workerThread()
                self.workerthread.signal.connect(self.errornormmsg)
                self.workerthread.start()
        except:
            self.workerthread = workerThread()
            self.workerthread.signal.connect(self.crashingmsg)
            self.workerthread.start()
            
    def clearClicked(self,checked=False):
        for edit in self.lineedits:
            if edit != self.saveedit[0]:
                edit.clear()
        for box in self.comboBox:
            box.setCurrentIndex(0)

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Message',
            "Are you sure to quit?", QMessageBox.Yes | 
            QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()
        

if __name__ == '__main__':
    
    app = QApplication(sys.argv)
    ex = General()
    sys.exit(app.exec_())
